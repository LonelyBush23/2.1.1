import csv
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side


data_entry = ["Введите название файла: ",
"Введите название профессии: "]


data_year = [lambda x:'Год',
lambda x:'Средняя зарплата',
lambda x:'Средняя зарплата - '+x,
lambda x:'Количество вакансий',
lambda x:'Количество вакансий - '+x]

data_sity = [lambda x:'Город', lambda x:'Уровень зарплат', lambda x:'', lambda x:'Город', lambda x:'Доля вакансий']

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,}

thins = Side(border_style="thin", color='000000')
style = Border(top=thins, bottom=thins, left=thins, right=thins)

class DataSet:
    def __init__(self, file_name: str):
        self.file_name = file_name
        self.list_vacancy = DataSet.create_list_vacancy(file_name)

    def сsv_reader(file_name : str):
        file = open(file_name, encoding='utf_8_sig')
        return csv.reader(file)

    def without_empty(my_list, count):
        try:
            my_list.remove('')
        except:
            pass
        return len(my_list) == count

    def create_list_vacancy(file_name : str):
        list_row = DataSet.сsv_reader(file_name)
        list_vacancy = []
        is_first_row = True
        for row in list_row:
            if is_first_row:
                naming_dic = DataSet.create_naming_dic(row)
                is_first_row = False
            elif DataSet.without_empty(row, len(naming_dic)):
                list_vacancy.append(Vacancy(row[naming_dic.get('name')], row[naming_dic.get('salary_from')], row[naming_dic.get('salary_to')], row[naming_dic.get('salary_currency')], row[naming_dic.get('area_name')], row[naming_dic.get('published_at')][:4]))
        return list_vacancy

    def create_naming_dic(naming):
        result = {}
        for i in range(len(naming)):
            result[naming[i]] = i
        return result

    def find_dynamics(self, profession):
        years_all_data = {}
        years_profession = {}
        sity_all_data = {}
        for vacancy in self.list_vacancy:
            years_all_data = DataSet.update(years_all_data, vacancy.published_at, vacancy.salary.current_salary)
            sity_all_data = DataSet.update(sity_all_data, vacancy.area_name, vacancy.salary.current_salary)
            if profession in vacancy.name and len(profession) != 0:
                years_profession = DataSet.update(years_profession, vacancy.published_at, vacancy.salary.current_salary)
            elif vacancy.published_at not in years_profession.keys():
                years_profession[vacancy.published_at] = (0, 0)
        return Report(profession, years_all_data, years_profession, sity_all_data)


    def update(dictionary, key, current_salary):
        if key in dictionary.keys():
            dictionary[key] = (dictionary[key][0] + current_salary, dictionary[key][1] + 1)
        else:
            dictionary[key] = (current_salary, 1)
        return dictionary


class Vacancy:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.salary = Salary(salary_from, salary_to, salary_currency)
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.current_salary = currency_to_rub[salary_currency] * (int(float(salary_to)) + int(float(salary_from))) // 2

class Report:
    def __init__(self, profession, years_all_data, years_profession, sity_all_data):
        self.profession = profession
        self.years = list(years_all_data.keys())
        self.sr_salary = Report.find_sal(years_all_data)
        self.salary_count = [v2 for v1,v2 in years_all_data.values()]
        self.sr_prof_salary = Report.find_sal(years_profession)
        self.prof_salary_count = [v2 for v1,v2 in years_profession.values()]
        self.all_vacancy_count = sum(self.salary_count)
        self.sity_for_salary, self.sity_salary, self.sity_for_vacancy, self.sity_vacancy = Report.find_sity_key(sity_all_data, self.all_vacancy_count)

    def find_sal(dictionary):
        sal = []
        for v1,v2 in dictionary.values():
            try:
                sal.append(int(float(v1 // v2)))
            except:
                sal.append(v1)
        return (sal)


    def find_sity_key(dictionary, all_vacancy_count):
        res_dic1 = {}
        res_dic2 = {}
        for key, value in dictionary.items():
            if all_vacancy_count / 100 <= value[1]:
                res_dic1[key] = int(float(value[0] // value[1]))
                res_dic2[key] = round(value[1] / all_vacancy_count, 4)

        res_dic1 = {k: v for k, v in sorted(res_dic1.items(), key=lambda item: item[1], reverse=True)}
        res_dic2 = {k: v for k, v in sorted(res_dic2.items(), key=lambda item: item[1], reverse=True)}

        res_dic1 = dict(list(res_dic1.items())[:10])
        res_dic2 = dict(list(res_dic2.items())[:10])
        return (list(res_dic1.keys()),list(res_dic1.values()), list(res_dic2.keys()),list(res_dic2.values()))


    def generate_excel(self):
        wb = Workbook()
        ws1 = wb.create_sheet('Статистика по годам')
        ws2 = wb.create_sheet('Статистика по городам')
        wb.remove(wb['Sheet'])
        for i in range(len(self.years)):
            self.fill_sheet(ws1, i, 1, self.years, data_year, int)
            self.fill_sheet(ws1, i, 2, self.sr_salary, data_year, int)
            self.fill_sheet(ws1, i, 3, self.sr_prof_salary, data_year, int)
            self.fill_sheet(ws1, i, 4, self.salary_count, data_year, int)
            self.fill_sheet(ws1, i, 5, self.prof_salary_count, data_year, int)
            if i < len(self.sity_for_salary):
                self.fill_sheet(ws2, i, 1, self.sity_for_salary, data_sity, str)
                self.fill_sheet(ws2, i, 2, self.sity_salary, data_sity, int)
                self.fill_sheet(ws2, i, 3, [''] * len(self.sity_for_salary), data_sity, str)
                self.fill_sheet(ws2, i, 4, self.sity_for_vacancy, data_sity, str)
                self.fill_sheet(ws2, i, 5, self.sity_vacancy, data_sity, float)
        for ws in wb:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value != '':
                        cell.border = style
        wb.save('report.xlsx')

    def fill_sheet(self, sheet, i, column, value, naming, format):
        if i == 0:
            sheet.cell(row=i + 1, column=column).value = naming[column-1](self.profession)
            sheet.cell(row=i + 1, column=column).font = Font(bold=True)
            my_value = [len(str(v)) for v in value]
            sheet.column_dimensions[get_column_letter(column)].width = max(len(naming[column - 1](self.profession)), max(my_value))+2
        sheet.cell(row=i + 2,column=column).value = format(value[i])
        if format == float:
            sheet.cell(row=i + 2, column=column).number_format = '0.00%'


""""*****************************************************************************************************************"""
information = []

for i in range(len(data_entry)):
    information.append(input(data_entry[i]))

dataSet = DataSet(information[0])
profession = information[1]

report = dataSet.find_dynamics(profession)
report.generate_excel()

